"""
mock_cache.py
In-memory implementation of LibraryCache for tests.
Loads directly from the locked Excel — no MongoDB or Redis needed.
Mirrors the production cache interface so route() runs unchanged.
"""
import re
from pathlib import Path
from openpyxl import load_workbook


def _parse_pmids(s):
    if not s or s == "see evidence base":
        return []
    return [x.strip() for x in re.split(r"\s*\|\s*", s) if x.strip()]


def _parse_zones(s):
    if not s: return []
    if str(s).lower() in ("multiple", "any"): return ["any"]
    return [z.strip() for z in str(s).split(",") if z.strip()]


def _label_to_band_key(label):
    """'Extreme Cold (<5°C)' → 'extreme_cold'. Used to parse guest cell band labels."""
    if not label: return None
    # Strip parenthetical range like '(<5°C)' or '(10–19%)'
    clean = re.sub(r"\s*\(.*?\)\s*", "", str(label)).strip().lower()
    return clean.replace(" ", "_")


class MockLibraryCache:
    """In-memory equivalent of the production LibraryCache."""

    LIBRARY_VERSION = "1.0.0"

    def __init__(self, xlsx_path):
        self.cells = {}            # (factor, band_key, skin_type, concern) → cell dict
        self.compound = {}         # (scenario_id, skin_type, concern) → cell dict
        self.guest = {}            # (cell_type, factor_or_scenario, skin_type) → cell dict
        self.scenarios = {}        # scenario_id → scenario dict
        self.age_mods = {}         # (age_band, concern) → mod dict
        self.gender_mods = {}      # (state, concern) → mod dict
        self.warmed = False
        self._load(xlsx_path)

    def _load(self, xlsx_path):
        wb = load_workbook(xlsx_path, data_only=True)

        # Single-factor cells from Master Library sheet (18 columns)
        # id|factor|band|range|band_key|skin|concern|risk|risk_label|
        # confidence|evidence|pmids|L0|L1|L2|action|zones|cities_season
        ws = wb["10. Master Library"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            cell = {
                "_id": row[0],
                "factor": row[1],
                "band_label": row[2],
                "range": row[3],
                "band_key": row[4],
                "skin_type": row[5],
                "concern": row[6],
                "risk": row[7],
                "risk_label": row[8],
                "confidence": row[9],
                "evidence_summary": row[10],
                "anchors": _parse_pmids(row[11]),
                "alerts": {"L0": row[12], "L1": row[13], "L2": row[14]},
                "action_cluster": row[15],
                "applicable_zones": _parse_zones(row[16]),
                "india_context": row[17],
            }
            self.cells[(row[1], row[4], row[5], row[6])] = cell

        # Compound cells
        ws = wb["9. Compound Cell Library"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            scen_id = row[0].split("-")[0]
            cell = {
                "_id": row[0], "scenario_id": scen_id,
                "scenario_name": row[1], "skin_type": row[2], "concern": row[3],
                "risk": row[4], "risk_label": row[5], "confidence": row[6],
                "evidence_summary": row[7],
                "alerts": {"L0": row[8], "L1": row[9], "L2": row[10]},
                "action_cluster": row[11],
                "zones": _parse_zones(row[12]),
                "seasons": row[13], "indian_cities": row[14],
            }
            self.compound[(scen_id, row[2], row[3])] = cell

        # Guest cells (single-factor + compound, 13 cols, data from row 5)
        # Key shape: (cell_type, factor_or_scenario, band_key_or_None, skin_type)
        # - single_factor cells: band_key from "Band / Drivers" label
        # - compound cells: band_key = None (scenario implies bands)
        ws = wb["11. Guest Mode"]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[0] or not str(row[0]).startswith("G-"): continue
            cell_type = "compound" if row[1] == "Compound" else "single_factor"
            band_key = _label_to_band_key(row[3]) if cell_type == "single_factor" else None
            cell = {
                "_id": row[0], "cell_type": cell_type,
                "factor_or_scenario": row[2], "band_or_drivers": row[3],
                "band_key": band_key,
                "skin_type": row[4], "concern": row[5],
                "risk": row[6], "risk_label": row[7],
                "alerts": {"L0": row[8], "L1": row[9], "L2": row[10]},
                "action_cluster": row[11],
                "zones_context": row[12],
            }
            self.guest[(cell_type, row[2], band_key, row[4])] = cell

        # Compound scenario index
        ws = wb["8. Compound Scenarios Index"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            self.scenarios[row[0]] = {
                "_id": row[0], "name": row[1],
                "bands": {"Temperature": row[2], "UV": row[3], "AQI": row[4], "Humidity": row[5]},
                "dominant_drivers": [d.strip() for d in (row[6] or "").split(",") if d.strip()],
                "zones": _parse_zones(row[7]),
                "seasons": [s.strip() for s in (row[8] or "").split(",") if s.strip()],
                "headline_mechanism": row[9],
                "cities_example": row[10],
            }

        # Age modifiers
        ws = wb["12. Age Modifiers"]
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Age Band" and ws.cell(row=r, column=3).value == "Risk Delta":
                for rr in range(r + 1, ws.max_row + 1):
                    ab = ws.cell(row=rr, column=1).value
                    cn = ws.cell(row=rr, column=2).value
                    if not ab or not cn: break
                    delta_str = str(ws.cell(row=rr, column=3).value or "0").replace("+", "")
                    self.age_mods[(ab, cn)] = {
                        "risk_delta": int(delta_str),
                        "addendum": ws.cell(row=rr, column=5).value,
                        "evidence": ws.cell(row=rr, column=6).value,
                    }
                break

        # Gender modifiers
        ws = wb["13. Gender + Life-Stage"]
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "State" and ws.cell(row=r, column=3).value == "Risk Delta":
                for rr in range(r + 1, ws.max_row + 1):
                    st = ws.cell(row=rr, column=1).value
                    cn = ws.cell(row=rr, column=2).value
                    if not st or not cn: break
                    delta_str = str(ws.cell(row=rr, column=3).value or "0").replace("+", "")
                    self.gender_mods[(st, cn)] = {
                        "risk_delta": int(delta_str),
                        "addendum": ws.cell(row=rr, column=5).value,
                        "evidence": ws.cell(row=rr, column=6).value,
                    }
                break

        self.warmed = True

    # ============================================================
    # ASYNC INTERFACE — mirrors production LibraryCache
    # ============================================================
    async def get_cell(self, factor, band_key, skin_type, concern):
        return self.cells.get((factor, band_key, skin_type, concern))

    async def get_compound_cell(self, scenario_id, skin_type, concern):
        return self.compound.get((scenario_id, skin_type, concern))

    async def get_guest_cell(self, cell_type, factor_or_scenario, skin_type, band_key=None):
        return self.guest.get((cell_type, factor_or_scenario, band_key, skin_type))

    # Diagnostic summary
    def summary(self):
        return {
            "cells": len(self.cells),
            "compound": len(self.compound),
            "guest": len(self.guest),
            "scenarios": len(self.scenarios),
            "age_mods": len(self.age_mods),
            "gender_mods": len(self.gender_mods),
        }
