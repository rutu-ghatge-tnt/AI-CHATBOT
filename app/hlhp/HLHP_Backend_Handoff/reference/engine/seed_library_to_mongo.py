"""
seed_library_to_mongo.py
Migrates the locked Excel library into MongoDB collections.
Run once per library version. Idempotent via library_version tagging.

Usage:
  pip install openpyxl pymongo
  export MONGO_URL="mongodb://localhost:27017"
  python seed_library_to_mongo.py --version 1.0.0 \
      --xlsx ./SkinBB_HLHP_Scenario_Library_v3_3.xlsx
"""
import argparse
import os
import re
from datetime import datetime, timezone
from openpyxl import load_workbook
from pymongo import MongoClient, ASCENDING


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True, help="Path to locked library .xlsx")
    p.add_argument("--version", required=True, help="Library version e.g. 1.0.0")
    p.add_argument("--db", default="skinbb_hlhp")
    p.add_argument("--mongo-url", default=os.getenv("MONGO_URL", "mongodb://localhost:27017"))
    p.add_argument("--dry-run", action="store_true", help="Validate without writing")
    return p.parse_args()


def parse_pmids(pmid_str):
    """Extract structured PMID/DOI list from the combined anchors string."""
    if not pmid_str or pmid_str == "see evidence base":
        return []
    return [s.strip() for s in re.split(r"\s*\|\s*", pmid_str) if s.strip()]


def parse_zones(zone_str):
    """Convert 'CN, HD, TP' or 'multiple' to a list."""
    if not zone_str:
        return []
    if zone_str.lower() in ("multiple", "any"):
        return ["any"]
    return [z.strip() for z in zone_str.split(",") if z.strip()]


# ============================================================
# COLLECTION BUILDERS — one function per collection
# ============================================================

def build_scenario_cells(wb, version):
    """Single-factor full-profile cells from Master Library sheet.

    Master Library has 18 columns (A..R) and NO 'points' column:
    0 ID | 1 Factor | 2 Band | 3 Range | 4 Band Key | 5 Skin Type | 6 Concern |
    7 Risk | 8 Risk Level | 9 Confidence | 10 Evidence | 11 PMID | 12 L0 |
    13 L1 | 14 L2 | 15 Action | 16 Zones | 17 India Cities/Season.
    Mapping kept in lock-step with mock_cache.MockLibraryCache.
    """
    docs = []
    ws = wb["10. Master Library"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        docs.append({
            "_id": row[0],                              # Scenario ID
            "factor": row[1],
            "band_label": row[2],
            "range": row[3],
            "band_key": row[4],
            "skin_type": row[5],
            "concern": row[6],
            "risk": row[7],
            "risk_label": row[8],
            "confidence": row[9],
            "evidence": {
                "summary": row[10],
                "anchors": parse_pmids(row[11]),
            },
            "alerts": {
                "L0": row[12],
                "L1": row[13],
                "L2": row[14],
            },
            "action_cluster": row[15],
            "applicable_zones": parse_zones(row[16]),
            "india_context": row[17],
            "library_version": version,
            "created_at": datetime.now(timezone.utc),
        })
    return docs


def build_compound_cells(wb, version):
    docs = []
    ws = wb["9. Compound Cell Library"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        docs.append({
            "_id": row[0],
            "scenario_id": row[0].split("-")[0],        # e.g. "C06" from "C06-OIL-ACNE-0001"
            "scenario_name": row[1],
            "skin_type": row[2],
            "concern": row[3],
            "risk": row[4],
            "risk_label": row[5],
            "confidence": row[6],
            "evidence_summary": row[7],
            "alerts": {
                "L0": row[8],
                "L1": row[9],
                "L2": row[10],
            },
            "action_cluster": row[11],
            "zones": parse_zones(row[12]),
            "seasons": row[13],
            "indian_cities": row[14],
            "library_version": version,
            "created_at": datetime.now(timezone.utc),
        })
    return docs


def build_guest_cells(wb, version):
    docs = []
    ws = wb["11. Guest Mode"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0] or not row[0].startswith("G-"):
            continue
        is_compound = row[1] == "Compound"
        # Derive band_key from label for single-factor cells (needed for lookup)
        band_key = None
        if not is_compound and row[3]:
            clean = re.sub(r"\s*\(.*?\)\s*", "", str(row[3])).strip().lower()
            band_key = clean.replace(" ", "_")
        docs.append({
            "_id": row[0],
            "cell_type": "compound" if is_compound else "single_factor",
            "factor_or_scenario": row[2],
            "band_or_drivers": row[3],
            "band_key": band_key,
            "skin_type": row[4],
            "concern": row[5],                          # "None" for guest cells
            "risk": row[6],
            "risk_label": row[7],
            "alerts": {
                "L0": row[8],
                "L1": row[9],
                "L2": row[10],
            },
            "action_cluster": row[11],
            "zones_context": row[12],
            "library_version": version,
            "created_at": datetime.now(timezone.utc),
        })
    return docs


def build_compound_scenarios(wb, version):
    docs = []
    ws = wb["8. Compound Scenarios Index"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        docs.append({
            "_id": row[0],                              # C01..C21
            "name": row[1],
            "bands": {
                "Temperature": row[2],
                "UV": row[3],
                "AQI": row[4],
                "Humidity": row[5],
            },
            "dominant_drivers": [d.strip() for d in (row[6] or "").split(",") if d.strip()],
            "zones": parse_zones(row[7]),
            "seasons": [s.strip() for s in (row[8] or "").split(",") if s.strip()],
            "headline_mechanism": row[9],
            "cities_example": row[10],
            "library_version": version,
        })
    return docs


def build_age_modifiers(wb, version):
    """Age × concern modifier rules. Found in '12. Age Modifiers' sheet."""
    docs = []
    ws = wb["12. Age Modifiers"]
    # find the modifier table by scanning for the header row
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Age Band" and ws.cell(row=r, column=3).value == "Risk Delta":
            for rr in range(r + 1, ws.max_row + 1):
                age_band = ws.cell(row=rr, column=1).value
                concern = ws.cell(row=rr, column=2).value
                if not age_band or not concern:
                    break
                delta = ws.cell(row=rr, column=3).value
                docs.append({
                    "_id": f"AGE-{age_band[:3].upper()}-{concern.split()[0][:4].upper()}",
                    "age_band": age_band,
                    "concern": concern,
                    "risk_delta": int(str(delta).replace("+", "")) if delta else 0,
                    "direction": ws.cell(row=rr, column=4).value,
                    "addendum": ws.cell(row=rr, column=5).value,
                    "evidence": ws.cell(row=rr, column=6).value,
                    "library_version": version,
                })
            break
    return docs


def build_gender_modifiers(wb, version):
    docs = []
    ws = wb["13. Gender + Life-Stage"]
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "State" and ws.cell(row=r, column=3).value == "Risk Delta":
            for rr in range(r + 1, ws.max_row + 1):
                state = ws.cell(row=rr, column=1).value
                concern = ws.cell(row=rr, column=2).value
                if not state or not concern:
                    break
                delta = ws.cell(row=rr, column=3).value
                docs.append({
                    "_id": f"GEN-{state.replace(' + ', '_').replace(' ', '')[:12]}-{concern.split()[0][:4].upper()}",
                    "state": state,
                    "concern": concern,
                    "risk_delta": int(str(delta).replace("+", "")) if delta else 0,
                    "direction": ws.cell(row=rr, column=4).value,
                    "addendum": ws.cell(row=rr, column=5).value,
                    "evidence": ws.cell(row=rr, column=6).value,
                    "library_version": version,
                })
            break
    return docs


def build_zones(wb, version):
    docs = []
    ws = wb["1. India Climatic Zones"]
    for row in ws.iter_rows(min_row=2, max_row=7, values_only=True):
        if not row[0] or row[0] in ("Zone Code",):
            continue
        docs.append({
            "_id": row[0],                              # HD, HH, CN, TP, CH, TN
            "name": row[1],
            "description": row[2],
            "dominant_skin_stressors": [s.strip() for s in (row[3] or "").split("·") if s.strip()],
            "cities_sample": row[4],
            "city_count": row[5],
            "library_version": version,
        })
    return docs


def build_band_definitions(wb, version):
    docs = []
    ws = wb["2. Bands Reference"]
    current_factor = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            current_factor = row[0]
            continue
        if not row[1]:
            continue
        docs.append({
            "_id": f"{current_factor}-{row[4]}",
            "factor": current_factor,
            "band_label": row[1],
            "range": row[2],
            "points": row[3],
            "band_key": row[4],
            "library_version": version,
        })
    return docs


# ============================================================
# INDEXES — built for the routing engine's lookup patterns
# ============================================================

def create_indexes(db):
    db.scenario_cells.create_index([
        ("library_version", ASCENDING),
        ("factor", ASCENDING),
        ("band_key", ASCENDING),
        ("skin_type", ASCENDING),
        ("concern", ASCENDING),
    ], unique=True, name="cell_lookup")

    db.compound_cells.create_index([
        ("library_version", ASCENDING),
        ("scenario_id", ASCENDING),
        ("skin_type", ASCENDING),
        ("concern", ASCENDING),
    ], unique=True, name="compound_lookup")

    db.guest_cells.create_index([
        ("library_version", ASCENDING),
        ("cell_type", ASCENDING),
        ("factor_or_scenario", ASCENDING),
        ("band_key", ASCENDING),
        ("skin_type", ASCENDING),
    ], name="guest_lookup")

    db.age_modifiers.create_index([
        ("library_version", ASCENDING),
        ("age_band", ASCENDING),
        ("concern", ASCENDING),
    ], unique=True, name="age_lookup")

    db.gender_modifiers.create_index([
        ("library_version", ASCENDING),
        ("state", ASCENDING),
        ("concern", ASCENDING),
    ], unique=True, name="gender_lookup")


# ============================================================
# MAIN
# ============================================================

def main():
    args = parse_args()
    print(f"Loading workbook: {args.xlsx}")
    wb = load_workbook(args.xlsx, data_only=True)
    version = args.version

    collections = {
        "scenario_cells":      build_scenario_cells(wb, version),
        "compound_cells":      build_compound_cells(wb, version),
        "guest_cells":         build_guest_cells(wb, version),
        "compound_scenarios":  build_compound_scenarios(wb, version),
        "age_modifiers":       build_age_modifiers(wb, version),
        "gender_modifiers":    build_gender_modifiers(wb, version),
        "zones":               build_zones(wb, version),
        "band_definitions":    build_band_definitions(wb, version),
    }

    print(f"\nLibrary version: {version}")
    print("Document counts per collection:")
    total = 0
    for name, docs in collections.items():
        print(f"  {name:.<30}{len(docs):>5}")
        total += len(docs)
    print(f"  {'TOTAL':.<30}{total:>5}")

    if args.dry_run:
        print("\n[DRY RUN] Validation complete. No writes performed.")
        return

    client = MongoClient(args.mongo_url)
    db = client[args.db]
    print(f"\nWriting to MongoDB: {args.mongo_url} → db={args.db}")
    for name, docs in collections.items():
        if not docs:
            continue
        # Idempotent: delete this version first, then insert
        db[name].delete_many({"library_version": version})
        db[name].insert_many(docs)
        print(f"  Inserted {len(docs)} docs into {name}")

    create_indexes(db)
    print("\nIndexes created.")
    print(f"\n✓ Library v{version} seeded. {total} documents across 8 collections.")


if __name__ == "__main__":
    main()
