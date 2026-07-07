"""Read SkinBB_HLHP_Scenario_Library workbook into a JSON-serialisable snapshot."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_XLSX = (
    Path(__file__).resolve().parents[1] / "data" / "SkinBB_HLHP_Scenario_Library_v3.5.xlsx"
)


def norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", norm(s).lower()).strip("_")


def band_key_for_label(bands: dict[str, list[dict[str, Any]]], factor: str, label: str) -> str:
    """Resolve a workbook band label to its band key using the Bands Reference sheet."""
    label = norm(label)
    for row in bands.get(factor, []):
        row_label = norm(row.get("label", ""))
        if label == row_label:
            return norm(row.get("key", "")) or slug(row_label)
        if label and row_label and label.split("(")[0].strip().lower() == row_label.split("(")[0].strip().lower():
            return norm(row.get("key", "")) or slug(row_label)
    return slug(label.split("(")[0] if "(" in label else label)


def _sheet_rows(wb: openpyxl.Workbook, name: str) -> list[list[Any]]:
    ws = wb[name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _library_version(path: Path) -> str:
    name = path.name.lower()
    if "v3.5" in name or "v3_5" in name:
        return "3.5"
    return "3.5"


def build_scenario_snapshot(xlsx_path: Path | str = DEFAULT_XLSX) -> dict[str, Any]:
    """Flatten the scenario library workbook into one snapshot dict."""
    path = Path(xlsx_path)
    version = _library_version(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = set(wb.sheetnames)

    bands: dict[str, list[dict[str, Any]]] = {}
    cur: str | None = None
    for r in _sheet_rows(wb, "2. Bands Reference")[1:]:
        factor, label, rng, pts, key = norm(r[0]), norm(r[1]), norm(r[2]), r[3], norm(r[4])
        if factor and not label:
            cur = factor
            bands[cur] = []
        elif cur and label:
            bands[cur].append(
                {
                    "label": label,
                    "range": rng,
                    "points": pts if isinstance(pts, (int, float)) else None,
                    "key": key,
                }
            )

    master: dict[str, dict[str, Any]] = {}
    rows = _sheet_rows(wb, "10. Master Library")
    for r in rows[1:]:
        if not norm(r[0]):
            continue
        factor, band, rng, bkey = norm(r[1]), norm(r[2]), norm(r[3]), norm(r[4])
        skin, concern = norm(r[5]), norm(r[6])
        cell = {
            "id": norm(r[0]),
            "factor": factor,
            "band": band,
            "range": rng,
            "band_key": bkey,
            "skin": skin,
            "concern": concern,
            "risk": r[7],
            "risk_level": norm(r[8]),
            "confidence": norm(r[9]),
            "evidence": norm(r[10]),
            "pmids": [p.strip() for p in norm(r[11]).split("|") if p.strip()],
            "l0": norm(r[12]),
            "l1": norm(r[13]),
            "l2": norm(r[14]),
            "action": norm(r[15]),
            "zones": [z.strip() for z in norm(r[16]).split(",") if z.strip()],
            "season": norm(r[17]),
        }
        master[f"{slug(factor)}|{bkey}|{slug(skin)}|{slug(concern)}"] = cell

    skins = sorted({c["skin"] for c in master.values() if c["skin"]})
    concerns = sorted({c["concern"] for c in master.values() if c["concern"]})
    factors = sorted({c["factor"] for c in master.values() if c["factor"]})

    compounds: list[dict[str, Any]] = []
    for r in _sheet_rows(wb, "8. Compound Scenarios Index")[1:]:
        if not norm(r[0]):
            continue
        compounds.append(
            {
                "id": norm(r[0]),
                "name": norm(r[1]),
                "temp_band": norm(r[2]),
                "uv_band": norm(r[3]),
                "aqi_band": norm(r[4]),
                "rh_band": norm(r[5]),
                "drivers": [d.strip() for d in norm(r[6]).split(",") if d.strip()],
                "zones": [z.strip() for z in norm(r[7]).split(",") if z.strip()],
                "seasons": norm(r[8]),
                "mechanism": norm(r[9]),
                "cities": norm(r[10]),
            }
        )

    compound_cells: dict[str, dict[str, Any]] = {}
    for r in _sheet_rows(wb, "9. Compound Cell Library")[1:]:
        if not norm(r[0]):
            continue
        name, skin, concern = norm(r[1]), norm(r[2]), norm(r[3])
        compound_cells[f"{slug(name)}|{slug(skin)}|{slug(concern)}"] = {
            "id": norm(r[0]),
            "scenario": name,
            "skin": skin,
            "concern": concern,
            "risk": r[4],
            "risk_level": norm(r[5]),
            "confidence": norm(r[6]),
            "evidence": norm(r[7]),
            "l0": norm(r[8]),
            "l1": norm(r[9]),
            "l2": norm(r[10]),
            "action": norm(r[11]),
        }

    zones: dict[str, dict[str, Any]] = {}
    city_zone: dict[str, str] = {}
    for r in _sheet_rows(wb, "1. India Climatic Zones")[1:]:
        code, name, desc, stressors, cities = norm(r[0]), norm(r[1]), norm(r[2]), norm(r[3]), norm(r[4])
        if not code:
            continue
        clist = [c.strip() for c in re.split(r"[;,]", cities) if c.strip()]
        zones[code] = {"code": code, "name": name, "desc": desc, "stressors": stressors, "cities": clist}
        for c in clist:
            city_zone[c.lower()] = code

    zone_weather = {
        "HH": {"temperature_c": 32, "aqi": 120, "uv_index": 7, "humidity_pct": 82},
        "CN": {"temperature_c": 34, "aqi": 210, "uv_index": 8, "humidity_pct": 38},
        "HD": {"temperature_c": 40, "aqi": 130, "uv_index": 10, "humidity_pct": 18},
        "TP": {"temperature_c": 28, "aqi": 80, "uv_index": 6, "humidity_pct": 52},
        "CH": {"temperature_c": 6, "aqi": 40, "uv_index": 7, "humidity_pct": 30},
        "TN": {"temperature_c": 29, "aqi": 60, "uv_index": 5, "humidity_pct": 90},
    }

    nuggets: list[dict[str, Any]] = []
    started = False
    for r in _sheet_rows(wb, "16. Did-You-Know Nuggets"):
        c0 = norm(r[0])
        if c0 == "#":
            started = True
            continue
        if started and c0 and c0.isdigit():
            nuggets.append({"n": int(c0), "text": norm(r[1]), "factor": norm(r[2]), "source": norm(r[3])})

    # ---- 11. Guest Mode: profile-less / no-concern cells -------------------
    guest: dict[str, dict[str, Any]] = {}
    guest_started = False
    for r in _sheet_rows(wb, "11. Guest Mode"):
        c0 = norm(r[0])
        if c0 == "Scenario ID":
            guest_started = True
            continue
        if not guest_started or not c0 or not c0.startswith("G-"):
            continue
        cell_type = norm(r[1])
        factor_or_name = norm(r[2])
        band_label = norm(r[3])
        skin = norm(r[4]) or "Normal"
        concern = norm(r[5]) or "None"
        cell = {
            "id": c0,
            "type": cell_type,
            "factor": factor_or_name if cell_type == "Single" else "",
            "scenario": factor_or_name if cell_type == "Compound" else "",
            "band": band_label,
            "band_key": (
                band_key_for_label(bands, factor_or_name, band_label)
                if cell_type == "Single"
                else ""
            ),
            "skin": skin,
            "concern": concern,
            "risk": r[6],
            "risk_level": norm(r[7]),
            "confidence": "Calibrated",
            "evidence": "",
            "pmids": [],
            "l0": norm(r[8]),
            "l1": norm(r[9]),
            "l2": norm(r[10]),
            "action": norm(r[11]),
            "zones": [z.strip() for z in norm(r[12]).split(",") if z.strip()],
            "season": "",
        }
        if cell_type == "Compound":
            key = f"compound|{slug(factor_or_name)}|{slug(skin)}|none"
        else:
            key = f"single|{slug(factor_or_name)}|{cell['band_key']}|{slug(skin)}|none"
        guest[key] = cell

    def modifiers(sheet_name: str) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        mod_started = False
        for r in _sheet_rows(wb, sheet_name):
            if norm(r[0]) == "Modifier":
                mod_started = True
                continue
            if mod_started and norm(r[0]):
                out.append(
                    {
                        "modifier": norm(r[0]),
                        "effect": norm(r[1]),
                        "concerns": norm(r[2]),
                        "direction": norm(r[3]),
                        "confidence": norm(r[4]),
                        "evidence": norm(r[5]),
                        "source": norm(r[6]),
                    }
                )
        return out

    gender_states: list[dict[str, Any]] = []
    gender_rules: dict[str, dict[str, Any]] = {}
    sheet_gender = "13. Gender + Life-Stage"
    if sheet_gender in sheet_names:
        sec: str | None = None
        for r in _sheet_rows(wb, sheet_gender):
            c0 = norm(r[0])
            if c0 == "State" and norm(r[1]) == "Description":
                sec = "ref"
                continue
            if c0 == "State" and norm(r[1]) == "Concern":
                sec = "rules"
                continue
            if not c0:
                continue
            if sec == "ref":
                if c0.startswith("State x Concern") or c0.startswith("State × Concern"):
                    sec = None
                    continue
                if c0 not in ("Gender + Life-Stage State Reference",) and norm(r[1]):
                    gender_states.append({"state": c0, "description": norm(r[1])})
            elif sec == "rules":
                concern = norm(r[1])
                if not concern:
                    continue
                delta_raw = norm(r[2]).replace("Δ", "").replace("+", "")
                try:
                    delta = int(float(delta_raw))
                except (TypeError, ValueError):
                    delta = 0
                gender_rules[f"{slug(c0)}|{slug(concern)}"] = {
                    "state": c0,
                    "concern": concern,
                    "risk_delta": delta,
                    "direction": norm(r[3]),
                    "action": norm(r[4]),
                    "addendum": norm(r[5]),
                    "anchor": norm(r[6]),
                }

    label_to_key: dict[tuple[str, str], str] = {}
    for fac, bl in bands.items():
        for b in bl:
            label_to_key[(fac, b["label"].lower())] = b["key"]
    time_overlay: dict[str, dict[str, str]] = {}
    sheet_time = "Time Overlay"
    if sheet_time in sheet_names:
        started = False
        for r in _sheet_rows(wb, sheet_time):
            if norm(r[0]) == "Type" and "Factor" in norm(r[1]):
                started = True
                continue
            fac = norm(r[0])
            if started and fac in ("UV", "AQI", "Temperature", "Humidity"):
                band_label = norm(r[2])
                key = label_to_key.get((fac, band_label.lower()))
                if not key:
                    continue
                morning = norm(r[3])
                if morning.startswith("—"):
                    morning = ""
                time_overlay[f"{slug(fac)}|{key}"] = {
                    "morning": morning,
                    "evening": norm(r[4]),
                }

    return {
        "meta": {
            "source": path.name,
            "source_path": str(path),
            "version": version,
            "note": (
                "Scenario library export — L0/L1/L2 alert text, risk, confidence, "
                "PMID anchors, gender/life-stage modifiers, time-of-day overlay. "
                "Product-free; 'advice' not used."
            ),
            "master_cell_count": len(master),
            "compound_cell_count": len(compound_cells),
            "guest_cell_count": len(guest),
            "gender_state_count": len(gender_states),
            "gender_rule_count": len(gender_rules),
            "time_overlay_count": len(time_overlay),
        },
        "bands": bands,
        "skins": skins,
        "concerns": concerns,
        "factors": factors,
        "zones": zones,
        "zone_weather": zone_weather,
        "city_zone": city_zone,
        "master": master,
        "compounds": compounds,
        "compound_cells": compound_cells,
        "guest": guest,
        "nuggets": nuggets,
        "nutrition": modifiers("14. Nutrition Modifiers"),
        "lifestyle": modifiers("15. Lifestyle Modifiers"),
        "gender_states": gender_states,
        "gender_rules": gender_rules,
        "time_overlay": time_overlay,
    }
