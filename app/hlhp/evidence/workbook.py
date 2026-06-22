"""Read all HLHP_Evidence_Base.xlsx sheets into JSON-serialisable structures."""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from app.hlhp.core.trigger_bands import normalize_rh_band, trigger_bands_snapshot
from app.hlhp.evidence.models import INTERNAL_SCIENCE_MARKER

FACTOR_SHEETS = [
    "UV",
    "Temperature",
    "Humidity",
    "Pollution",
    "Nutritional Status",
    "Lifestyle",
]

VALID_BANDS = {
    "season": {
        "any",
        "summer",
        "monsoon",
        "winter",
        "spring",
        "autumn",
        "winter_dry",
        "winter_humid",
        "pre_monsoon",
        "post_monsoon",
    },
    "uvi": {"any", "off", "low", "moderate", "high", "very_high", "extreme"},
    "aqi": {
        "any",
        "good",
        "satisfactory",
        "moderate",
        "poor",
        "very_poor",
        "severe",
    },
    "rh": {"any", "very_low", "low", "moderate", "comfortable", "high", "very_high"},
    "temp": {
        "any",
        "very_cold",
        "cold",
        "comfortable",
        "warm",
        "hot",
        "very_hot",
    },
}

SUNSCREEN_RE = re.compile(r"\b(sunscreen|spf|outdoor protection)\b", re.I)
_DAYTIME_UVI_BANDS = ["low", "moderate", "high", "very_high", "extreme"]


def mentions_sunscreen(text: str) -> bool:
    return bool(text and SUNSCREEN_RE.search(text))


def _parse_band_list(raw: object, *, dimension: str | None = None) -> list[str]:
    if raw is None:
        return ["any"]
    text = str(raw).strip().lower()
    if not text or text == "any":
        return ["any"]
    bands = [part.strip() for part in text.split(",") if part.strip()]
    if dimension == "rh":
        bands = [normalize_rh_band(b) for b in bands]
    return bands


def _col_index(headers: list[object], *prefixes: str) -> int | None:
    """Resolve column by exact name or prefix (handles mojibake in L2 header)."""
    normalized = {str(h).strip(): i for i, h in enumerate(headers) if h}
    for prefix in prefixes:
        if prefix in normalized:
            return normalized[prefix]
    for header, idx in normalized.items():
        for prefix in prefixes:
            if header.startswith(prefix):
                return idx
    return None


def _cell(excel_row: tuple, col: dict[str, int], *header_candidates: str) -> str:
    for candidate in header_candidates:
        for name, idx in col.items():
            if name == candidate or name.startswith(candidate):
                if idx < len(excel_row):
                    val = excel_row[idx]
                    return (val or "").strip() if val is not None else ""
    return ""


def _detect_internal_only(*texts: str) -> bool:
    marker = INTERNAL_SCIENCE_MARKER.lower()
    return any(marker in (t or "").lower() for t in texts)


def _parse_user_filter(raw: object) -> list[dict[str, str]]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text or text.lower() == "any":
        return []
    tokens: list[dict[str, str]] = []
    for part in text.split(","):
        part = part.strip()
        if not part or ":" not in part:
            continue
        cls, val = part.split(":", 1)
        tokens.append({"class": cls.strip().lower(), "value": val.strip().lower()})
    return tokens


def _source_citation(row: dict[str, object]) -> str:
    title = row.get("source_title") or ""
    pages = row.get("pages_doi_pmid") or ""
    chapter = row.get("chapter_section") or ""
    edition = row.get("edition_year") or ""
    bits = [str(b).strip() for b in (title, edition, chapter, pages) if b and str(b).strip()]
    return " — ".join(bits) if bits else "HLHP Evidence Base"


def autofix_night_gate_uvi(row: dict, autofixes: list[str]) -> None:
    uvi_bands = row["triggers"]["uvi"]
    if uvi_bands != ["any"]:
        return
    l1_guest = row.get("alert_l1_guest") or ""
    l1_personalised = row.get("alert_l1_personalised") or ""
    if not (mentions_sunscreen(l1_guest) or mentions_sunscreen(l1_personalised)):
        return
    row["triggers"]["uvi"] = list(_DAYTIME_UVI_BANDS)
    autofixes.append(
        f"{row['id']}: Trigger_UVI_Band auto-set to {', '.join(_DAYTIME_UVI_BANDS)} "
        "(sunscreen in L1; excludes off/night)"
    )


def read_factor_sheet(ws, factor: str) -> list[dict]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    rows: list[dict] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not excel_row or not excel_row[0]:
            continue
        row_num = excel_row[col["Row #"]]
        row_id = f"{factor[:3].upper()}-{int(row_num)}"
        user_tokens = _parse_user_filter(excel_row[col["Trigger_User_Filter"]])
        never_fire = any(
            t["class"] == "flag" and t["value"] == "never_fire" for t in user_tokens
        )
        user_tokens = [t for t in user_tokens if t["class"] != "flag"]

        row = {
            "id": row_id,
            "factor": factor,
            "row_number": int(row_num),
            "sub_effect": (excel_row[col["Sub-effect / metric"]] or "").strip(),
            "quantified": (excel_row[col["Quantified value / threshold"]] or "").strip(),
            "mechanism": (excel_row[col["Mechanism (1 line)"]] or "").strip(),
            "product_implication": (
                excel_row[col["Product / ingredient implication"]] or ""
            ).strip(),
            "outcome_tag": (excel_row[col["Outcome tag"]] or "").strip(),
            "confidence": (excel_row[col["Confidence"]] or "").strip(),
            "india_relevant": str(excel_row[col["India-relevant (Y/N)"]]).upper() == "Y",
            "source_type": (excel_row[col["Source Type"]] or "").strip(),
            "source_title": (excel_row[col["Source: Title (Book or Paper)"]] or "").strip(),
            "edition_year": (excel_row[col["Edition / Year"]] or "").strip(),
            "chapter_section": (
                excel_row[col["Chapter / Section / Journal+Author"]] or ""
            ).strip(),
            "pages_doi_pmid": (excel_row[col["Pages / DOI / PMID"]] or "").strip(),
            "alert_short": (excel_row[col["Alert_Short (consumer-facing)"]] or "").strip(),
            "priority": (excel_row[col["Trigger_Priority"]] or "P2").strip(),
            "triggers": {
                "season": _parse_band_list(excel_row[col["Trigger_Season"]]),
                "uvi": _parse_band_list(excel_row[col["Trigger_UVI_Band"]]),
                "aqi": _parse_band_list(excel_row[col["Trigger_AQI_Band"]]),
                "rh": _parse_band_list(excel_row[col["Trigger_RH_Band"]], dimension="rh"),
                "temp": _parse_band_list(excel_row[col["Trigger_Temp_Band"]]),
                "user_filter": user_tokens,
            },
            "alert_l1_personalised": (
                excel_row[col["Alert_L1_Personalised (rewritten)"]] or ""
            ).strip(),
            "alert_l1_guest": (excel_row[col["Alert_L1_Guest (no profile)"]] or "").strip(),
            "alert_l1_evening_personalised": _cell(
                excel_row, col, "Alert_L1_Evening_Personalised"
            ),
            "alert_l1_evening_guest": _cell(excel_row, col, "Alert_L1_Evening_Guest"),
            "alert_l2_explainer": _cell(excel_row, col, "Alert_L2_Explainer"),
            "time_of_day_phase": _cell(excel_row, col, "Time_of_Day_Phase") or "any_time",
            "mood_verdict_tag": _cell(excel_row, col, "Mood_Verdict_Tag"),
            "combination_stack": _cell(excel_row, col, "Combination_Stack"),
            "engagement_archetype": _cell(excel_row, col, "Engagement_Archetype"),
            "physical_analogy": _cell(excel_row, col, "Physical_Analogy"),
            "body_sensation_decode": _cell(excel_row, col, "Body_Sensation_Decode"),
            "symptom_keyword": _cell(excel_row, col, "Symptom_Keyword"),
            "routine_action": _cell(excel_row, col, "Routine_Action"),
            "visual_icon_hint": _cell(excel_row, col, "Visual_Icon_Hint"),
            "never_fire": never_fire,
            "science_citation": "",
        }
        row["internal_only"] = _detect_internal_only(
            row["alert_l1_personalised"],
            row["alert_l1_guest"],
            row["alert_l1_evening_personalised"],
            row["alert_l1_evening_guest"],
        )
        row["science_citation"] = _source_citation(row)
        rows.append(row)
    return rows


def read_science_nuggets(ws) -> list[dict]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    nuggets: list[dict] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not excel_row or not excel_row[0]:
            continue
        book = excel_row[col["Source: Book"]] or ""
        edition = excel_row[col["Edition"]] or ""
        page = excel_row[col["Page"]] or ""
        source = " — ".join(str(x).strip() for x in (book, edition, page) if x)
        nuggets.append(
            {
                "id": int(excel_row[col["#"]]),
                "text": (excel_row[col["Nugget (consumer-friendly, ≤30 words)"]] or "").strip(),
                "factor": (excel_row[col["Factor"]] or "").strip(),
                "source_type": (excel_row[col["Source Type"]] or "").strip(),
                "source": source or "HLHP Science Nuggets",
            }
        )
    return nuggets


def read_book_inventory(ws) -> list[dict]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    books: list[dict] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not excel_row or not excel_row[col["#"]]:
            continue
        books.append(
            {
                "id": int(excel_row[col["#"]]),
                "title": (excel_row[col["Title"]] or "").strip(),
                "author": (excel_row[col["Author / Editor"]] or "").strip(),
                "edition": (excel_row[col["Edition"]] or "").strip(),
                "focus": (excel_row[col["Apparent focus (one-line)"]] or "").strip(),
                "priority": (excel_row[col["Priority (P0/P1/P2)"]] or "").strip(),
                "india_relevant": str(excel_row[col["India-relevant (Y/N)"]]).upper() == "Y",
                "extraction_status": (excel_row[col["Extraction status"]] or "").strip(),
            }
        )
    return books


def read_glossary(ws) -> list[dict]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    entries: list[dict] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not excel_row or not excel_row[col["#"]]:
            continue
        entries.append(
            {
                "id": int(excel_row[col["#"]]),
                "category": (excel_row[col["Category"]] or "").strip(),
                "term": (excel_row[col["Jargon term"]] or "").strip(),
                "lay_translation": (excel_row[col["Lay translation"]] or "").strip(),
                "notes": (excel_row[col["Notes"]] or "").strip(),
            }
        )
    return entries


def read_gaps_conflicts(ws) -> list[dict]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    items: list[dict] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not excel_row or not excel_row[col["#"]]:
            continue
        items.append(
            {
                "id": int(excel_row[col["#"]]),
                "type": (excel_row[col["Type (Gap / Conflict)"]] or "").strip(),
                "topic": (excel_row[col["HLHP question or topic"]] or "").strip(),
                "value_a": excel_row[col["Value A"]],
                "source_a": (excel_row[col["Source A (book, edition, page)"]] or "").strip(),
                "value_b": excel_row[col["Value B"]],
                "source_b": (excel_row[col["Source B (book, edition, page)"]] or "").strip(),
                "note": (excel_row[col["Note"]] or "").strip(),
            }
        )
    return items


def read_readme(ws) -> dict[str, Any]:
    meta: dict[str, Any] = {"title": "", "sections": {}, "totals": {}}
    current_key: str | None = None
    for excel_row in ws.iter_rows(min_row=1, values_only=True):
        if not excel_row:
            continue
        a = excel_row[0]
        b = excel_row[1] if len(excel_row) > 1 else None
        if a and not b and isinstance(a, str) and "Evidence Base" in a:
            meta["title"] = a.strip()
            continue
        if a and b and isinstance(a, str) and isinstance(b, str):
            key = a.strip()
            meta["sections"][key] = b.strip()
            current_key = key
            continue
        if a and isinstance(a, str) and a.strip().startswith("Total:"):
            meta["totals"]["summary"] = a.strip()
        if a and isinstance(a, str) and "·" in a and any(
            f in a for f in ("UV", "Temperature", "Humidity")
        ):
            for part in a.split("·"):
                part = part.strip()
                if ":" in part:
                    k, v = part.split(":", 1)
                    meta["totals"][k.strip()] = v.strip()
    return meta


def read_coverage_matrix(ws) -> dict[str, Any]:
    """Parse the multi-grid Coverage_Matrix sheet."""
    grids: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for excel_row in ws.iter_rows(min_row=1, values_only=True):
        if not excel_row or not excel_row[0]:
            continue
        label = str(excel_row[0]).strip()
        if label.startswith("Trigger ×"):
            if current:
                grids.append(current)
            current = {"title": label, "columns": [], "rows": []}
            continue
        if current is None:
            continue
        if label == "Trigger":
            current["columns"] = [str(c).strip() for c in excel_row[1:] if c is not None]
            continue
        if label in {
            "UV",
            "Temperature",
            "Humidity",
            "Pollution",
            "Sleep",
            "Stress",
            "Nutritional Status",
            "Lifestyle",
        }:
            values = []
            for c in excel_row[1 : 1 + len(current["columns"])]:
                try:
                    values.append(int(c) if c is not None else 0)
                except (TypeError, ValueError):
                    values.append(0)
            current["rows"].append({"trigger": label, "counts": values})
    if current:
        grids.append(current)
    return {"grids": grids, "legend": {"gap": 0, "thin": 2, "adequate": 3}}


def build_snapshot(xlsx_path: Path) -> dict[str, Any]:
    from app.hlhp.evidence.citations import validate_citations
    from app.hlhp.evidence.coverage import (
        build_coverage_report,
        compute_coverage_grids,
        fill_coverage_gaps,
        write_coverage_matrix_sheet,
    )
    from app.hlhp.evidence.index import build_inverted_index
    from app.hlhp.evidence.voice import apply_lay_voice, sanitize_l1_percentages, validate_l1_voice

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    findings: list[dict] = []
    for sheet in FACTOR_SHEETS:
        findings.extend(read_factor_sheet(wb[sheet], sheet))

    snapshot = {
        "version": 3,
        "workbook_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_workbook": xlsx_path.name,
        "trigger_bands": trigger_bands_snapshot(),
        "readme": read_readme(wb["README"]),
        "book_inventory": read_book_inventory(wb["Book Inventory"]),
        "finding_count": len(findings),
        "findings": findings,
        "science_nuggets": read_science_nuggets(wb["Science Nuggets"]),
        "nugget_count": 0,
        "glossary": read_glossary(wb["Glossary"]),
        "gaps_conflicts": read_gaps_conflicts(wb["Gaps & Conflicts"]),
        "coverage_matrix": read_coverage_matrix(wb["Coverage_Matrix"]),
    }
    from app.hlhp.evidence.composition_workbook import read_all_composition

    snapshot["composition"] = read_all_composition(wb)
    snapshot["composition_row_count"] = sum(
        v for k, v in snapshot["composition"].get("_counts", {}).items()
    )
    snapshot["nugget_count"] = len(snapshot["science_nuggets"])
    wb.close()

    coverage_fixes = fill_coverage_gaps(findings)
    computed_grids = compute_coverage_grids(findings)
    snapshot["coverage_matrix"] = {
        "grids": computed_grids,
        "legend": {"gap": 0, "thin": 2, "adequate": 3},
        "source": "auto-generated",
    }
    snapshot["coverage_tag_fixes"] = coverage_fixes

    autofixes: list[str] = [f"{f['row_id']}: {f['detail']}" for f in coverage_fixes]
    warnings: list[str] = []
    glossary = snapshot["glossary"]
    l1_fields = (
        "alert_l1_guest",
        "alert_l1_personalised",
        "alert_l1_evening_guest",
        "alert_l1_evening_personalised",
    )
    for row in findings:
        for field in l1_fields:
            text = row.get(field) or ""
            if not text:
                continue
            cleaned = sanitize_l1_percentages(text)
            cleaned = apply_lay_voice(cleaned, glossary)
            if cleaned != text:
                row[field] = cleaned
                autofixes.append(f"{row['id']}: sanitised {field}")
        row["internal_only"] = _detect_internal_only(
            *(row.get(f) or "" for f in l1_fields)
        )
        autofix_night_gate_uvi(row, autofixes)
        for band_key, values in row["triggers"].items():
            if band_key == "user_filter":
                continue
            for band in values:
                allowed = VALID_BANDS.get(band_key, {"any"})
                if band not in allowed:
                    warnings.append(f"{row['id']}: invalid {band_key} band '{band}'")
        uvi_bands = row["triggers"]["uvi"]
        if uvi_bands == ["any"]:
            for text in (row.get("alert_l1_guest") or "", row.get("alert_l1_personalised") or ""):
                if mentions_sunscreen(text):
                    warnings.append(
                        f"{row['id']}: sunscreen L1 with Trigger_UVI_Band=any (night-gate risk)"
                    )
                    break
        if not row.get("source_title") or not row.get("pages_doi_pmid"):
            warnings.append(f"{row['id']}: incomplete citation")

    from app.hlhp.evidence.validation_gates import validate_snapshot_v2

    voice_issues = validate_l1_voice(findings, glossary)
    citation_issues = validate_citations(findings, snapshot["book_inventory"])
    gate_failures = validate_snapshot_v2(
        findings, glossary, snapshot["book_inventory"]
    )
    coverage_report = build_coverage_report(findings)
    snapshot["inverted_index"] = build_inverted_index(findings)

    snapshot["build_autofixes"] = autofixes
    snapshot["build_warnings"] = warnings
    snapshot["build_report"] = {
        "voice_violations": voice_issues,
        "citation_issues": citation_issues,
        "gate_failures": gate_failures,
        "gate_failure_count": len(gate_failures),
        "coverage_report": coverage_report,
        "gaps_conflicts_count": len(snapshot["gaps_conflicts"]),
        "glossary_entries": len(snapshot["glossary"]),
        "book_inventory_count": len(snapshot["book_inventory"]),
    }
    return snapshot


def sync_workbook_coverage_matrix(xlsx_path: Path, grids: list[dict]) -> None:
    """Persist computed coverage grids back to the xlsx authoring sheet."""
    import openpyxl

    from app.hlhp.evidence.coverage import write_coverage_matrix_sheet

    wb = openpyxl.load_workbook(xlsx_path)
    write_coverage_matrix_sheet(wb["Coverage_Matrix"], grids)
    wb.save(xlsx_path)


def sync_workbook_user_filters(xlsx_path: Path, findings: list[dict[str, Any]]) -> int:
    """Write Trigger_User_Filter from snapshot findings back to factor sheets."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    updated = 0
    by_factor: dict[str, list[dict]] = defaultdict(list)
    for row in findings:
        by_factor[row["factor"]].append(row)

    for sheet in FACTOR_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        if "Trigger_User_Filter" not in headers:
            continue
        row_col = headers.index("Row #") + 1
        filter_col = headers.index("Trigger_User_Filter") + 1
        lookup = {int(r["row_number"]): r for r in by_factor.get(sheet, [])}
        for excel_row in range(2, ws.max_row + 1):
            row_num = ws.cell(excel_row, row_col).value
            if row_num is None:
                continue
            data = lookup.get(int(row_num))
            if not data:
                continue
            new_val = _format_user_filter_from_row(data)
            if ws.cell(excel_row, filter_col).value != new_val:
                ws.cell(excel_row, filter_col).value = new_val
                updated += 1
    wb.save(xlsx_path)
    return updated


def _format_user_filter_from_row(row: dict[str, Any]) -> str:
    tokens = row.get("triggers", {}).get("user_filter", [])
    if not tokens:
        return "any"
    return ", ".join(f"{t['class']}:{t['value']}" for t in tokens)
