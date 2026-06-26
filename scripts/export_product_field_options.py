"""Export product form field option lists from skin_bb MongoDB (read-only) to a simple Excel/CSV."""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Any

from pymongo import MongoClient

DEFAULT_MONGO_URI = "mongodb://skinbb:IgR9RMn%40skNBbh84Gsz@localhost:27018/skin_bb"
DB_NAME = "skin_bb"
OUTPUT_DIR = Path(__file__).resolve().parents[1] / "app" / "label_looker" / "data"

_OBJECT_ID_RE = re.compile(r"^[0-9a-f]{24}$", re.I)


def _display_name(doc: dict[str, Any]) -> str:
    for key in ("label", "name", "title"):
        v = doc.get(key)
        if isinstance(v, str) and v.strip():
            return v.strip()
    v = doc.get("value") or doc.get("slug")
    return str(v).strip() if v else ""


def _is_active(doc: dict[str, Any]) -> bool:
    if doc.get("isDeleted") is True:
        return False
    if doc.get("isActive") is False:
        return False
    return True


def _is_garbage_benefit(doc: dict[str, Any]) -> bool:
    value = str(doc.get("value") or "").strip()
    name = _display_name(doc)
    if not value or not name:
        return True
    if _OBJECT_ID_RE.fullmatch(value) or _OBJECT_ID_RE.fullmatch(name):
        return True
    if value == name and re.fullmatch(r"[0-9a-f]{20,}", value):
        return True
    return False


def _options_from_collection(db, collection: str, *, active_only: bool = True) -> list[str]:
    names: list[str] = []
    for doc in db[collection].find({}):
        if active_only and not _is_active(doc):
            continue
        name = _display_name(doc)
        if name:
            names.append(name)
    return sorted(set(names), key=str.lower)


def _options_from_attribute_slug(db, slug: str) -> list[str]:
    attr = db.product_attributes.find_one({"slug": slug, "isDeleted": {"$ne": True}})
    if not attr:
        return []
    names: list[str] = []
    for doc in db.product_attribute_values.find(
        {"attributeId": attr["_id"], "isDeleted": {"$ne": True}, "isActive": {"$ne": False}}
    ):
        name = _display_name(doc)
        if name:
            names.append(name)
    return sorted(set(names), key=str.lower)


def _benefit_options(db) -> list[str]:
    names: list[str] = []
    for doc in db.benefits.find({}):
        if _is_garbage_benefit(doc):
            continue
        name = _display_name(doc)
        if name:
            names.append(name)
    return sorted(set(names), key=str.lower)


def _dedupe_sorted(names: list[str]) -> list[str]:
    return sorted(set(names), key=str.lower)


def _merged_options(*sources: list[str]) -> list[str]:
    combined: list[str] = []
    for src in sources:
        combined.extend(src)
    return _dedupe_sorted(combined)


def _certification_options(db) -> list[str]:
    from_attr = _options_from_attribute_slug(db, "certification-applicable")
    from_coll = _options_from_collection(db, "certification_applicables")
    return _merged_options(from_attr, from_coll)


def build_field_options(db) -> list[tuple[str, list[str]]]:
    """Return (field_name, options) groups in UI order."""
    return [
        ("Benefits", _benefit_options(db)),
        ("Product classification", _options_from_attribute_slug(db, "product-classification")),
        ("Body part", _options_from_attribute_slug(db, "body-part")),
        ("Makeup finish", _options_from_attribute_slug(db, "makeup-finish")),
        ("Fragrance family", _options_from_attribute_slug(db, "fragrance-family")),
        ("Special feature", _options_from_attribute_slug(db, "special-feature")),
        ("Certifications", _certification_options(db)),
        ("Formulation", _merged_options(
            _options_from_collection(db, "formulations"),
            _options_from_attribute_slug(db, "formulation"),
        )),
        ("Gender", _merged_options(
            _options_from_collection(db, "product_genders"),
            _options_from_collection(db, "genders"),
            _options_from_attribute_slug(db, "gender"),
        )),
        ("Claims", _options_from_attribute_slug(db, "claims")),
        ("Tags", _options_from_collection(db, "product_tags")),
        ("Hair type", _options_from_collection(db, "product_hair_types")),
        ("Skin type", _options_from_collection(db, "product_skin_types")),
        ("Skin concerns", _options_from_collection(db, "product_skin_concerns")),
        ("Hair concerns", _options_from_collection(db, "product_hair_concerns")),
        ("Hair goals", _options_from_collection(db, "hair_goals")),
        ("Conscious", _options_from_collection(db, "conscious")),
        ("Meta keywords", _options_from_collection(db, "product_tags")),
    ]


def _write_csv(path: Path, groups: list[tuple[str, list[str]]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Field", "Option"])
        for field_name, options in groups:
            for i, option in enumerate(options):
                writer.writerow([field_name if i == 0 else "", option])


def _write_excel(path: Path, groups: list[tuple[str, list[str]]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise SystemExit("openpyxl required: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Field Options"
    ws.append(["Field", "Option"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    row_idx = 2
    for field_name, options in groups:
        if not options:
            ws.append([field_name, ""])
            row_idx += 1
            continue
        start_row = row_idx
        for option in options:
            ws.append(["", option])
            row_idx += 1
        end_row = row_idx - 1
        ws.cell(row=start_row, column=1, value=field_name)
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export simple field → option list (read-only).")
    parser.add_argument("--mongo-uri", default=DEFAULT_MONGO_URI)
    parser.add_argument("--format", choices=["csv", "xlsx", "both"], default="both")
    args = parser.parse_args()

    client = MongoClient(args.mongo_uri, serverSelectionTimeoutMS=10000)
    groups = build_field_options(client[DB_NAME])
    client.close()

    total_options = sum(len(opts) for _, opts in groups)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path = OUTPUT_DIR / "product_field_options.csv"
    xlsx_path = OUTPUT_DIR / "product_field_options.xlsx"

    if args.format in ("csv", "both"):
        _write_csv(csv_path, groups)
        print(f"CSV: {csv_path} ({total_options} options)")

    if args.format in ("xlsx", "both"):
        _write_excel(xlsx_path, groups)
        print(f"Excel: {xlsx_path} ({total_options} options)")

    print("\nOptions per field:")
    for name, options in groups:
        print(f"  {name}: {len(options)}")


if __name__ == "__main__":
    main()
